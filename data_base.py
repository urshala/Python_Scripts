import openpyxl
import os





def _db_exists(db_name):
	'''Returns true if the given excel file exists '''
	if os.path.exists(db_name):
		return True 
	return False


def create_db(db_name, list_ot_headers):
	''' Creates the excel sheet and with the given parameters as headers of excel
	file '''


	def insert_value():
		for item in list_ot_headers:
			yield item
	table_generator = insert_value()
	
	db_name = db_name + '.xlsx'
	if _db_exists(db_name):
		raise Exception('The database with that name already exits!')
	if not isinstance(list_ot_headers, list):
		raise Exception(' Make sure that the headers is in list format')

	excel_file = openpyxl.Workbook()
	active_sheet = excel_file.active
	number_of_columns = len(list_ot_headers)
	for index in range(1, number_of_columns + 1):
		active_sheet.cell(row=1, column=index).value = next(table_generator)

	excel_file.save(db_name)
	print 'Database {} created successfully'.format(db_name)

	


def insert_into_db(db_name, list_of_value):
	''' Saves the given list of values into excel sheet '''
	db = db_name
	db_name = db_name + '.xlsx'
	if not _db_exists(db_name):
		raise Exception('Database with name "{}" does not exist'.format(db))
	wb = openpyxl.load_workbook(db_name)
	active_sheet = wb.active
	column = active_sheet.max_column
	if not isinstance(list_of_value, list):
		raise Exception('Enter the data to be inserted as a list')
	my_gen = (item for item in list_of_value)

	if len(list_of_value) % column != 0:
		raise Exception('Make sure the data has correct length')

	outer_loop_counter = len(list_of_value)/ column
	for counter in range(outer_loop_counter):
		row = active_sheet.max_row
		for i in range(1, column+1):
			active_sheet.cell(row=row+1, column=i).value = next(my_gen)

	wb.save(db_name)
	print 'Successfully inserted data'





def delete_from_db(db_name, list_of_value, field='first_column'):
	''' Deletes the record with matching value and field 
		E.g. delete_from_db('MyDB', 'first_name', 'First Name' '''
	db = db_name
	db_name = db_name + '.xlsx'

	if not _db_exists(db_name):
		raise Exception('Database with name "{}" does not exist'.format(db))

	wb = openpyxl.load_workbook(db_name)
	active_sheet = wb.active
	headers = show_tables(db_name)
	if field in headers:
		column_value = headers.index(field) + 1 
		row, col = active_sheet.max_row, active_sheet.max_column
		match_results = []
		for i in range(1,row+1):
			if active_sheet.cell(row=i, column=column_value).value == list_of_value:
				match_results.append((i,column_value))
		if len(match_results) > 1:
			print '''There are more than one match data in database. Which one do you want to delete?
			Press the required number. '''
			col_into_letter = '{}'.format(openpyxl.utils.get_column_letter(col))
			match_column_slice = ''
			for index,i in enumerate(match_results, start=1):
				s = ''
				match_column_slice = 'A{}:{}{}'.format(i[0], col_into_letter, i[0] )
				for match in active_sheet[match_column_slice]:
					for item in match:
						s = s + ' ' + str(item.value)
				print '{}. {}'.format(index, s)

			try:
				prm = input('Enter the number to delete item >')
				if not 1 <= prm <= len(match_results):
					raise Exception('Enter a valid number')
			except (NameError, TypeError):
				raise Exception('Enter a valid number')

			list_index_to_delete = match_results[prm - 1 ]
			column_slice_to_delete = list_index_to_delete[0]
			slice_val = 'A{0}:{1}{0}'. format(column_slice_to_delete, col_into_letter)
			for item in active_sheet[slice_val]:
				for i in item:
					i.value = None
			print 'Successfully deleted !!'

		elif len(match_results) == 1:			
			print 'Found one match'
			for i in match_results:
				match_column_slice = 'A{0}:{1}{0}'.format(i[0], col_into_letter)
				for match in active_sheet[match_column_slice]:
					for each_tup in match:
						each_tup.value = None			
			print 'Deleted value'	
		else:
			print '"{}" does not match any record '.format(list_of_value)		

	else:
		print '"{}" is not a valid column name'.format(field)

	wb.save(db_name)



def update_from_db(db_name, old_data, data_to_be_updated):
	''' Updates the last matching item '''
	db = db_name
	db_name = db_name + '.xlsx'
	if not _db_exists(db_name):
		raise Exception('Database with name "{}" does not exist'.format(db))
 

	wb = openpyxl.load_workbook(db_name)
	active_sheet = wb.active
	row, col = active_sheet.max_row, active_sheet.max_column
	old_data = old_data.split(',')
	od = ' '.join(old_data)
	print 'The od is ', od
	col_into_letter = '{}'.format(openpyxl.utils.get_column_letter(col))
	
	new_data = data_to_be_updated.split(',')
	
	replace_new_data = (item for item in new_data)

	if not len(old_data) == len(new_data) == col:
		raise Exception('Make sure the data length is correct.')
	match_slice = ''
	for r in range(2,row+1):
		s = ''
		for c in range(1,col+1):
			val = active_sheet.cell(row=r, column=c).value
			if val is None:
				continue
			s = s + ' ' + active_sheet.cell(row=r, column=c).value
		
		if s.strip() == od.strip():
			match_slice = 'A{0}:{1}{0}'.format(r,col_into_letter)
			
	for item in active_sheet[match_slice]:
		for each_tup in item:
			each_tup.value = next(replace_new_data)
			print each_tup.value
	print 'Successfully updated data'	
			
	wb.save(db_name)



def search_data(db_name, search_keywords, field):
	db = db_name
	db_name = db_name + '.xlsx'
	if not _db_exists(db_name):
		raise Exception('Database with name "{}" does not exist'.format(db))

	wb = openpyxl.load_workbook(db_name)
	active_sheet = wb.active
	headers = show_tables(db_name)
	if field in headers:
		column_value = headers.index(field) + 1 
		row, col = active_sheet.max_row, active_sheet.max_column
		match_results = []
		for i in range(1,row+1):
			if active_sheet.cell(row=i, column=column_value).value == search_keywords:
				match_results.append((i,column_value))
		
		print ' Your search for {} gives this result'.format(search_keywords)
		col_into_letter = '{}'.format(openpyxl.utils.get_column_letter(col))
		match_column_slice = ''
		for index,i in enumerate(match_results, start=1):
			s = ''
			match_column_slice = 'A{}:{}{}'.format(i[0], col_into_letter, i[0] )
			for match in active_sheet[match_column_slice]:
				for item in match:
					s = s + ' ' + item.value
			print '{}. {}'.format(index, s)

def show_db():
	data_bases = []
	for files in os.listdir('.'):
		if files.endswith('xlsx'):
			filename, _ = files.split('.')
			data_bases.append(filename)
	if len(data_bases) != 0:
		print 'Found following database\n'
		for item in data_bases:
			print item
	else:
		print 'Found no database'
	



def show_tables(db_name):
	if not db_name.endswith('.xlsx'):
		db_name = db_name + '.xlsx'
	else:
		db_name = db_name

	db = db_name
	if not _db_exists(db_name):
		raise Exception('Database with name "{}" does not exist'.format(db))
	wb = openpyxl.load_workbook(db_name)
	sheet = wb.active
	number_of_columns = sheet.max_column
	headers = []
	#print 'The tables in {} are as follows: '.format(db_name)
	for i in range(1, number_of_columns + 1):
		column_head = sheet.cell(row=1, column=i).value
		#print '{}. {}'.format(i,column_head)
		headers.append(column_head)
	

	return headers

def show_all_data(db_name):
	db = db_name
	db_name = db_name + '.xlsx'
	if not _db_exists(db_name):
		raise Exception('Database with name "{}" does not exist'.format(db))
	wb = openpyxl.load_workbook(db_name)
	sheet = wb.active
	row, col = sheet.max_row, sheet.max_column

	for r in range(2, row +1):
		result = ''
		for c in range(1,col +1):
			record = sheet.cell(row=r, column=c).value
			if record is None:
				continue
			result = result + ' ' + record
		if result:
			print result,
			print 



if '__name__' == '__main__':
	main()



#search_data('Myfile.xlsx', 'Jenga', 'First Name')
#delete_from_db('Myfile.xlsx', 'Jenga', 'First Name')
#show_tables('Myfile.xlsx')
#create_db('Myfile', ['First Name', 'Last Name', 'Age', 'Score'])
#insert_into_db('Myfile.xlsx', ['deepak', 'am', 'Yo-kyla', 'ED', 'Jenga', 'Koila', 12,'deepak'])
#show_db()